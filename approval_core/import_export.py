"""
Import/Export utilities for SMVS Approval System
Handles CSV and XLSX file processing + django-import-export resources
"""
import csv
import io
from django.contrib.auth.models import Group
from django.http import HttpResponse
from openpyxl import Workbook
from import_export import resources, fields
from import_export.fields import Field
from import_export.widgets import ManyToManyWidget, ForeignKeyWidget
from .models import ApprovalRule, User, UserRole, ApprovalLevelUser
from approval_core.models import (
    EmailNotificationTemplate, SMSTemplate, WhatsAppNotificationTemplate, Department, Country,
    Center, ReportPermission, RuleApprovalSequence, UserWorkspace, RegistrationWorkflowConfig, NotificationRoutingMatrix, ApprovalLevel,
    UserProfile
)


class ImportExportHelper:
    """Helper class for custom import/export operations"""

    @staticmethod
    def generate_csv_response(filename, headers, data_rows):
        """Generate CSV file response"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        writer = csv.writer(response)
        writer.writerow(headers)
        for row in data_rows:
            writer.writerow(row)
        
        return response
    
    @staticmethod
    def generate_xlsx_response(filename, sheet_name, headers, data_rows):
        """Generate XLSX file response"""
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name
        
        # Add headers with bold formatting
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = cell.font.copy()
            if hasattr(cell.font, 'bold'):
                cell.font.bold = True
        
        # Add data rows
        for row_num, row_data in enumerate(data_rows, 2):
            for col_num, value in enumerate(row_data, 1):
                worksheet.cell(row=row_num, column=col_num).value = value
        
        # Adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        workbook.save(response)
        return response
    
    @staticmethod
    def parse_csv_file(csv_file):
        """Parse CSV file and return data"""
        try:
            decoded_file = csv_file.read().decode('utf-8').splitlines()
            reader = csv.DictReader(decoded_file)
            return list(reader)
        except Exception as e:
            return None, str(e)
    
    @staticmethod
    def parse_xlsx_file(xlsx_file):
        """Parse XLSX file and return data"""
        try:
            from openpyxl import load_workbook
            workbook = load_workbook(xlsx_file)
            worksheet = workbook.active
            
            headers = []
            data = []
            
            for row_idx, row in enumerate(worksheet.iter_rows(values_only=True), 1):
                if row_idx == 1:
                    headers = [cell for cell in row if cell is not None]
                else:
                    if any(row):  # Skip empty rows
                        row_dict = dict(zip(headers, row))
                        data.append(row_dict)
            
            return data
        except Exception as e:
            return None, str(e)


# ==================== RESOURCES FOR django-import-export ====================

class ApprovalRuleResource(resources.ModelResource):
    """Upgraded Resource for ApprovalRule handling M2M relationships and Foreign Keys cleanly"""
    
    # 🟢 1. Department Condition (ManyToManyField) - Joins via department code (e.g., D_FIN, D_HR)
    applicable_departments = fields.Field(
        attribute='applicable_departments',
        column_name='applicable_departments',
        widget=ManyToManyWidget(Department, field='code', separator=',')
    )
    
    # 🟢 2. Applicable Country (ForeignKey) - Looks up country code dynamically (e.g., IN, US)
    applicable_country = fields.Field(
        attribute='applicable_country',
        column_name='applicable_country',
        widget=ForeignKeyWidget(Country, field='code')
    )
    
    # 🟢 3. Approval Levels (ManyToManyField with a 'through' table) - Explicitly marked for Export Only
    approval_levels = fields.Field(
        column_name='approval_levels',
        readonly=True  # 👈 Crucial safety flag: Prevents intermediate database lock conflicts during import
    )

    class Meta:
        model = ApprovalRule
        # 🟢 4. Include all requested parameters in your spreadsheet matrix layout
        fields = (
            'id', 
            'rule_name', 
            'rule_type', 
            'min_amount', 
            'max_amount', 
            'applicable_departments',  
            'approval_levels',          # 👈 Included safely in the fields array
            'applicable_country',      
            'chain_type',              
            'priority', 
            'is_active'
        )
        export_order = fields

    def dehydrate_approval_levels(self, approval_rule):
        """Custom exporter: Grabs all sequence levels for this rule and formats them as a text string (e.g., '1,2,3')"""
        # Fetch levels using the sequence order mapping setup
        sequences = approval_rule.rule_approval_sequences.all().select_related('approval_level')
        levels_list = [str(seq.approval_level.level_number) for seq in sequences if seq.approval_level]
        return ",".join(levels_list)

class UserResource(resources.ModelResource):
    """Upgraded Resource to export core fields alongside relational profile parameters"""
    
    center = fields.Field(column_name='Center')
    department = fields.Field(column_name='Department')
    mobile = fields.Field(column_name='Mobile Number')

    class Meta:
        model = User
        # 🟢 Include all personal info, permissions statuses, and relational profile metrics
        fields = (
            'id', 
            'username', 
            'email', 
            'first_name', 
            'last_name', 
            'mobile', 
            'center', 
            'department', 
            'is_active', 
            'is_staff', 
            'is_superuser', 
            'date_joined', 
            'last_login'
        )
        export_order = fields

    def dehydrate_mobile(self, user):
        """Safely fetch profile mobile fallback"""
        profile = getattr(user, 'user_profile', None)
        return profile.phone if profile and profile.phone else "-"

    def dehydrate_center(self, user):
        """Safely fetch profile center location name fallback"""
        profile = getattr(user, 'user_profile', None)
        return profile.center.name if profile and profile.center and profile.center.name else "-"

    def dehydrate_department(self, user):
        """Safely fetch profile department track name fallback"""
        profile = getattr(user, 'user_profile', None)
        return profile.department.name if profile and profile.department and profile.department.name else "-"


class UserRoleResource(resources.ModelResource):
    """Upgraded Resource for UserRole using matched key properties to clear double names."""
    
    # 🟢 Matching the variable names to the column names completely eliminates the parenthetical text
    id = fields.Field(attribute='id', column_name='Id')
    user_username = fields.Field(attribute='user__username', column_name='User Username')
    role = fields.Field(attribute='role', column_name='Role')
    department_name = fields.Field(attribute='department__name', column_name='Department Name')
    center_name = fields.Field(attribute='center__name', column_name='Center Name')
    mobile_number = fields.Field(attribute='mobile_number', column_name='Mobile Number')
    is_active = fields.Field(attribute='is_active', column_name='Is Active')
    access_rights = fields.Field(column_name='Access Rights', readonly=True)
   
    class Meta:
        model = UserRole
        # 🟢 Match these keys exactly to the clean keys mapped above
        fields = (
            'id', 
            'user_username', 
            'role', 
            'department_name', 
            'center_name', 
            'mobile_number', 
            'access_rights', 
            'is_active'
        )
        export_order = fields

    def dehydrate_access_rights(self, user_role):
        """Generates dynamic string format (e.g. 'C:105 D:49')"""
        try:
            centers_count = user_role.accessible_centers.count()
            depts_count = user_role.accessible_departments.count()
            return f"C:{centers_count} D:{depts_count}"
        except Exception:
            return "C:0 D:0"


class EmailMappingResource(resources.ModelResource):
    """Resource to allow full import/export matching for Email and WhatsApp Mappings"""
    class Meta:
        from .models import EmailMapping  # Imported locally to prevent circular locks
        model = EmailMapping
        # 🟢 ADDED: Added 'phone_number' explicitly into the engine fields mapping fields
        fields = ('id', 'post__role_name', 'mapping_type', 'center__code', 'department__code', 'email', 'phone_number', 'person_name', 'is_primary', 'is_active')
        export_order = fields

class GroupResource(resources.ModelResource):
    class Meta:
        model = Group
        fields = ('id', 'name')
        export_order = fields

class EmailNotificationTemplateResource(resources.ModelResource):
    class Meta:
        model = EmailNotificationTemplate
        fields = ('id', 'template_name', 'event_type', 'subject', 'body', 'approval_link_text', 'is_active')
        export_order = fields

class SMSTemplateResource(resources.ModelResource):
    class Meta:
        model = SMSTemplate
        fields = ('id', 'template_name', 'event_type', 'approval_level__level_name', 'message_text', 'is_active')
        export_order = fields

class WhatsAppNotificationTemplateResource(resources.ModelResource):
    class Meta:
        model = WhatsAppNotificationTemplate
        fields = ('id', 'template_name', 'event_type', 'message_body', 'button_text', 'button_url', 'is_active')
        export_order = fields

class ReportPermissionResource(resources.ModelResource):
    class Meta:
        model = ReportPermission
        fields = ('id', 'user__username', 'can_view_report', 'can_enter_actual_amount', 'can_enter_work_completion')
        export_order = fields

class RuleApprovalSequenceResource(resources.ModelResource):
    rule = fields.Field(
        column_name='rule_name',
        attribute='rule',
        widget=ForeignKeyWidget(ApprovalRule, field='rule_name')
    )
    approval_level = fields.Field(
        column_name='approval_level_name',
        attribute='approval_level',
        widget=ForeignKeyWidget(ApprovalLevel, field='level_name')
    )

    class Meta:
        model = RuleApprovalSequence
        fields = ('id', 'rule', 'approval_level', 'sequence_order', 'is_mandatory', 'allow_delegation', 'auto_approve_if_conditions_met')
        export_order = fields

class UserWorkspaceResource(resources.ModelResource):
    class Meta:
        model = UserWorkspace
        fields = ('id', 'user__username', 'departments__code')
        export_order = fields

class NotificationRoutingMatrixResource(resources.ModelResource):
    """Upgraded Resource matching your final Unified Notification Router layout schema configuration"""
    class Meta:
        model = NotificationRoutingMatrix
        # 🟢 UPDATED: This list matches your clean model properties perfectly with zero name mismatches
        fields = (
            'id', 
            'event_type', 
            
            # Email fields matching model variables
            'email_center_accountant', 'email_center', 'email_center_sant', 'email_prabhari_sant', 'email_zonal_head',
            'email_department', 'email_dept_leader_sant', 'email_dept_sant', 'email_hod', 'email_mk_haribhakt', 'email_mk_sant', 'email_secretary', 'email_end_user',
            
            # WhatsApp fields matching model variables
            'wa_center_accountant', 'wa_center', 'wa_center_sant', 'wa_prabhari_sant', 'wa_zonal_head',
            'wa_department', 'wa_dept_leader_sant', 'wa_dept_sant', 'wa_hod', 'wa_mk_haribhakt', 'wa_mk_sant', 'wa_secretary', 'wa_end_user',
            
            # SMS fields matching model variables
            'sms_center_accountant', 'sms_center', 'sms_center_sant', 'sms_prabhari_sant', 'sms_zonal_head',
            'sms_department', 'sms_dept_leader_sant', 'sms_dept_sant', 'sms_hod', 'sms_mk_haribhakt', 'sms_mk_sant', 'sms_secretary', 'sms_end_user'
        )
        export_order = fields


class ApprovalLevelResource(resources.ModelResource):
    class Meta:
        model = ApprovalLevel
        fields = ('id', 'level_number', 'level_name', 'description', 'is_active', 'show_time_taken')
        export_order = fields


class RegistrationWorkflowConfigResource(resources.ModelResource):
    class Meta:
        model = RegistrationWorkflowConfig
        fields = ('id', 'config_name', 'is_direct_registration', 'is_active', 'updated_at')
        export_order = fields


class ApprovalLevelUserResource(resources.ModelResource):
    """Resource to allow seamless import/export for Level Assignments with clear text names"""
    
    # 🟢 Explicitly map the fields so they match both the database attributes and your headers
    user = fields.Field(
        column_name='user',
        attribute='user',
        widget=ForeignKeyWidget(User, field='username')
    )
    
    approval_level = fields.Field(
        column_name='level',
        attribute='approval_level',  # Matches your model's ForeignKey relation field name
        widget=ForeignKeyWidget(ApprovalLevel, field='level_number') # Exports the level number (e.g., 1, 2)
    )
    
    departments = fields.Field(
        column_name='departments',
        attribute='departments',
        widget=ManyToManyWidget(Department, field='code', separator=',')
    )

    class Meta:
        model = ApprovalLevelUser
        # 🟢 These match your fields list exactly
        fields = ('id', 'user', 'approval_level', 'departments', 'is_active')
        export_order = fields

class UserProfileResource(resources.ModelResource):
    """Resource to map User Profile properties along with core authentication accounts"""
    user = fields.Field(attribute='user__username', column_name='user_username', widget=ForeignKeyWidget(User, field='username'))
    department = fields.Field(attribute='department__code', column_name='department_code', widget=ForeignKeyWidget(Department, field='code'))
    center = fields.Field(attribute='center__code', column_name='center_code', widget=ForeignKeyWidget(Center, field='code'))

    class Meta:
        model = UserProfile
        fields = ('id', 'user', 'phone', 'department', 'center')
        export_order = fields

# ==================== SAMPLE DATA FOR DOWNLOADS ====================

COUNTRIES_SAMPLE = {
    'headers': ['code', 'name', 'description'],
    'rows': [
        ['IN', 'India', 'Indian Subcontinent'],
        ['US', 'USA', 'United States'],
    ]
}

ZONES_SAMPLE = {
    'headers': ['country_code', 'code', 'name', 'description'],
    'rows': [
        ['IN', 'Z_NORTH', 'North India', 'Northern Region'],
        ['IN', 'Z_SOUTH', 'South India', 'Southern Region'],
        ['IN', 'Z_EAST', 'East India', 'Eastern Region'],
    ]
}

CENTERS_SAMPLE = {
    'headers': ['country_code', 'zone_code', 'code', 'name', 'city', 'state', 'pincode', 'address'],
    'rows': [
        ['IN', 'Z_NORTH', 'C_DEL', 'Delhi Center', 'New Delhi', 'Delhi', '110001', 'Main Street, New Delhi'],
        ['IN', 'Z_SOUTH', 'C_BNG', 'Bangalore Center', 'Bangalore', 'Karnataka', '560001', 'IT Hub, Bangalore'],
        ['IN', 'Z_EAST', 'C_KOL', 'Kolkata Center', 'Kolkata', 'West Bengal', '700001', 'Business District, Kolkata'],
    ]
}

DEPARTMENTS_SAMPLE = {
    'headers': ['country_code', 'code', 'name', 'description', 'center_code'],
    'rows': [
        ['IN', 'D_FIN', 'Finance', 'Finance Department', 'C_DEL'],
        ['IN', 'D_HR', 'HR', 'Human Resources', 'C_BNG'],
        ['IN', 'D_OPS', 'Operations', 'Operations Department', 'C_KOL'],
        ['IN', 'D_IT', 'IT', 'Information Technology', ''],
    ]
}

POSTS_SAMPLE = {
    'headers': ['post_type', 'role_name', 'description'],
    'rows': [
        ['center', 'Center Manager', 'Manager of the center'],
        ['center', 'Center Email', 'Email point of center'],
        ['department', 'Department Head', 'Head of department'],
        ['department', 'Department Email', 'Email point of department'],
    ]
}

EMAIL_MAPPINGS_SAMPLE = {
    # 🟢 UPDATED: Appended 'phone_number' column header text right after 'email'
    'headers': ['post_type', 'role_name', 'mapping_type', 'email', 'phone_number', 'person_name', 'center_code', 'department_code', 'is_primary'],
    'rows': [
        # 🟢 UPDATED: Included placeholder mobile strings mapping to the new column index positions
        ['center', 'Center Manager', 'center', 'manager@center.com', '918670058009', 'John Doe', 'C_DEL', '', 'Yes'],
        ['center', 'Center Email', 'center', 'email@center.com', '918670058009', 'Jane Smith', 'C_BNG', '', 'No'],
        ['department', 'Department Head', 'department', 'head@dept.com', '918670058009', 'Bob Wilson', '', 'D_FIN', 'Yes'],
        ['department', 'Department Email', 'department', 'email@dept.com', '918670058009', 'Alice Brown', '', 'D_HR', 'No'],
    ]
}

GROUPS_SAMPLE = {
    'headers': ['name'],
    'rows': [['System Auditors'], ['HOD Approvers']]
}

EMAIL_TEMPLATES_SAMPLE = {
    'headers': ['template_name', 'event_type', 'subject', 'body', 'approval_link_text', 'is_active'],
    'rows': [
        ['[WKF-01] Pending Verification Alert', 'pending_approval', 'Action Required: Form {form_number}', 'Hello {{ user.get_full_name }}, Form {form_number} is pending.', 'Click here to review', 'Yes']
    ]
}

SMS_TEMPLATES_SAMPLE = {
    'headers': ['template_name', 'event_type', 'approval_level_id', 'message_text', 'is_active'],
    'rows': [
        ['[WKF-01] Backup Text Notice', 'pending_approval', '', 'Form {form_number} is pending your approval level.', 'Yes']
    ]
}

WHATSAPP_TEMPLATES_SAMPLE = {
    'headers': ['template_name', 'event_type', 'message_body', 'button_text', 'button_url', 'is_active'],
    'rows': [
        ['[FRM-01] Submission Success Card', 'form_submitted', 'Jai Swaminarayan. Your form *{form_number}* for {source_name} was received.', 'View Form', 'https://harivadan.in', 'Yes']
    ]
}

REPORT_PERMISSIONS_SAMPLE = {
    'headers': ['username', 'can_view_report', 'can_enter_actual_amount', 'can_enter_work_completion'],
    'rows': [['hspatel', 'Yes', 'No', 'Yes']]
}

RULE_SEQUENCES_SAMPLE = {
    'headers': ['rule_name', 'approval_level_number', 'sequence_order', 'is_mandatory', 'allow_delegation', 'auto_approve_if_conditions_met'],
    'rows': [['Combined High Value Rule', '2', '1', 'Yes', 'No', 'No']]
}

USER_WORKSPACES_SAMPLE = {
    'headers': ['username', 'department_codes_separated_by_comma'],
    'rows': [['hspatel', 'D_FIN, D_HR']]
}

USER_ROLES_SAMPLE = {
    'headers': ['User Username', 'Role', 'Department Name', 'Center', 'Mobile Number', 'Access Rights', 'Is Active'],
    'rows': [
        ['nirav', 'MK Sabhya', 'BNIT - BN-IT', '', '+919925242806', 'C:0 D:4', 'Yes'],
        ['center_user', 'End User', '', 'VSN - Vasna', '+919998469102', 'C:0 D:1', 'Yes']
    ]
}

USERS_SAMPLE = {
    'headers': ['username', 'email', 'first_name', 'last_name', 'is_active'],
    'rows': [
        ['harshadpatel', 'harshad.patel@smvs.org', 'Harshad', 'Patel', 'Yes'],
        ['niravbasheri', 'nirav.basheri@smvs.org', 'Nirav', 'Basheri', 'Yes']
    ]
}

APPROVAL_LEVEL_USERS_SAMPLE = {
    'headers': ['user_username', 'approval_level_number', 'departments_by_code', 'is_active'],
    'rows': [
        ['harshadpatel', '1', 'D_FIN,D_HR', 'Yes'],
        ['niravbasheri', '2', 'D_OPS', 'Yes']
    ]
}

USER_PROFILES_SAMPLE = {
    'headers': ['user_username', 'phone', 'department_code', 'center_code'],
    'rows': [
        ['harshadpatel', '919925242806', 'D_FIN', ''],
        ['niravbasheri', '919998469102', '', 'C_DEL']
    ]
}